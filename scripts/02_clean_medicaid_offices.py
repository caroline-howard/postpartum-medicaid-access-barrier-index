"""Clean the raw Medicaid office locations workbook.

Input:
    data/raw/medicaid_offices.xlsx

Outputs:
    data/processed/medicaid_offices_clean.csv
    outputs/cleaning_summary.md

The raw Excel file is never modified. This script performs only the first
office-location cleaning step; it does not add county joins, ACS indicators,
CMS enrollment data, rurality fields, Power BI files, or index calculations.
"""

from __future__ import annotations

import csv
import hashlib
import re
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit(
        "Missing required package: openpyxl. Install it before running this script."
    ) from exc


PROJECT_ROOT = Path(__file__).resolve().parents[1]
RAW_INPUT = PROJECT_ROOT / "data" / "raw" / "medicaid_offices.xlsx"
CLEAN_OUTPUT = PROJECT_ROOT / "data" / "processed" / "medicaid_offices_clean.csv"
SUMMARY_OUTPUT = PROJECT_ROOT / "outputs" / "cleaning_summary.md"
ARTICLE_REPORTED_COUNT = 3026

EXPECTED_FIELDS = [
    "state_fips",
    "state_name",
    "agency_name",
    "street1",
    "street2",
    "city",
    "state",
    "zip_code",
    "latitude",
    "longitude",
]
REQUIRED_OFFICE_FIELDS = ["agency_name", "street1", "city", "state", "zip_code"]
SUSPICIOUS_TEXT_PATTERN = re.compile(
    r"\b(note|metadata|source|total|count)\b",
    flags=re.IGNORECASE,
)


def snake_case(value: Any) -> str:
    """Convert a spreadsheet column label to snake_case."""
    text = "" if value is None else str(value).strip().lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def clean_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def build_full_address(row: dict[str, str]) -> str:
    parts = [
        row.get("street1", ""),
        row.get("street2", ""),
        row.get("city", ""),
        row.get("state", ""),
        row.get("zip_code", ""),
    ]
    return ", ".join(part for part in parts if part)


def build_office_id(row: dict[str, str]) -> str:
    stable_values = [
        row.get("state_fips", ""),
        row.get("state_name", ""),
        row.get("agency_name", ""),
        row.get("full_address", ""),
        row.get("latitude", ""),
        row.get("longitude", ""),
    ]
    digest = hashlib.sha1("|".join(stable_values).encode("utf-8")).hexdigest()[:12]
    return f"office_{digest}"


def read_workbook(path: Path) -> tuple[list[str], list[dict[str, str]], list[str], str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet_names = workbook.sheetnames
    selected_sheet = sheet_names[0]
    worksheet = workbook[selected_sheet]
    rows = worksheet.iter_rows(values_only=True)

    try:
        raw_headers = next(rows)
    except StopIteration as exc:
        raise ValueError(f"Workbook is empty: {path}") from exc

    headers = [snake_case(header) for header in raw_headers]
    if len(headers) != len(set(headers)):
        duplicates = [header for header, count in Counter(headers).items() if count > 1]
        raise ValueError(f"Duplicate column names after snake_case conversion: {duplicates}")

    records: list[dict[str, str]] = []
    for excel_row_number, values in enumerate(rows, start=2):
        record = {
            header: clean_cell(value)
            for header, value in zip(headers, values, strict=False)
            if header
        }
        record["_excel_row_number"] = str(excel_row_number)
        records.append(record)

    return headers, records, sheet_names, selected_sheet


def validate_expected_fields(headers: list[str]) -> None:
    missing = [field for field in EXPECTED_FIELDS if field not in headers]
    if missing:
        expected = ", ".join(EXPECTED_FIELDS)
        found = ", ".join(headers)
        missing_text = ", ".join(missing)
        raise ValueError(
            "Missing expected field(s): "
            f"{missing_text}\nExpected fields: {expected}\nFound fields: {found}"
        )


def write_clean_csv(rows: list[dict[str, str]], source_headers: list[str]) -> None:
    CLEAN_OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = ["office_id", *source_headers, "full_address"]

    with CLEAN_OUTPUT.open("w", newline="", encoding="utf-8") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def parse_float(value: str) -> float | None:
    try:
        return float(value)
    except ValueError:
        return None


def summarize_row_checks(rows: list[dict[str, str]]) -> dict[str, Any]:
    missing_required = {
        field: sum(1 for row in rows if not row.get(field, "")) for field in REQUIRED_OFFICE_FIELDS
    }
    blank_rows = sum(
        1
        for row in rows
        if not any(row.get(field, "") for field in EXPECTED_FIELDS)
    )

    suspicious_rows: list[str] = []
    invalid_coordinate_rows = 0
    for row in rows:
        row_number = row.get("_excel_row_number", "unknown")
        state_name = row.get("state_name", "")
        agency_name = row.get("agency_name", "")
        latitude = row.get("latitude", "")
        longitude = row.get("longitude", "")

        reasons = []
        combined_text = f"{state_name} {agency_name}"
        if SUSPICIOUS_TEXT_PATTERN.search(combined_text):
            reasons.append("metadata-like text")

        latitude_value = parse_float(latitude)
        longitude_value = parse_float(longitude)
        if latitude_value is None or longitude_value is None:
            reasons.append("non-numeric coordinate")
            invalid_coordinate_rows += 1
        elif not (18 <= latitude_value <= 72 and -180 <= longitude_value <= -60):
            reasons.append("coordinate outside broad U.S. range")
            invalid_coordinate_rows += 1

        if reasons:
            suspicious_rows.append(f"Excel row {row_number}: {', '.join(reasons)}")

    duplicate_full_rows = sum(
        count - 1
        for count in Counter(
            tuple(row.get(field, "") for field in EXPECTED_FIELDS) for row in rows
        ).values()
        if count > 1
    )

    return {
        "missing_required": missing_required,
        "blank_rows": blank_rows,
        "invalid_coordinate_rows": invalid_coordinate_rows,
        "suspicious_rows": suspicious_rows,
        "duplicate_full_rows": duplicate_full_rows,
    }


def format_missing_required(missing_required: dict[str, int]) -> str:
    return "\n".join(
        f"- Rows missing `{field}`: {count}" for field, count in missing_required.items()
    )


def format_suspicious_rows(suspicious_rows: list[str]) -> str:
    if not suspicious_rows:
        return "- None found"
    rows = suspicious_rows[:10]
    if len(suspicious_rows) > 10:
        rows.append(f"...and {len(suspicious_rows) - 10} more")
    return "\n".join(f"- {row}" for row in rows)


def write_summary(
    raw_count: int,
    clean_count: int,
    removed_missing_coordinates: int,
    states_represented: int,
    duplicate_coordinate_pairs: int,
    sheet_names: list[str],
    selected_sheet: str,
    row_checks: dict[str, Any],
) -> None:
    SUMMARY_OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    run_time = datetime.now().astimezone().isoformat(timespec="seconds")
    sheet_note = (
        "The workbook contains one sheet, so `Sheet1` is the selected data sheet."
        if len(sheet_names) == 1
        else "The first sheet is selected; review additional sheets before changing this."
    )
    missing_required_text = format_missing_required(row_checks["missing_required"])
    suspicious_rows_text = format_suspicious_rows(row_checks["suspicious_rows"])
    if (
        row_checks["blank_rows"] == 0
        and row_checks["duplicate_full_rows"] == 0
        and row_checks["invalid_coordinate_rows"] == 0
        and not row_checks["suspicious_rows"]
        and all(count == 0 for count in row_checks["missing_required"].values())
    ):
        discrepancy_note = (
            f"The Data in Brief article reports {ARTICLE_REPORTED_COUNT:,} office "
            f"locations. The downloaded Excel file currently contains {clean_count:,} "
            "rows after cleaning. No metadata, blank, duplicated, or invalid row was "
            "identified by the current validation checks, so the cleaning step preserves "
            "the source file as downloaded and documents the count discrepancy "
            "transparently."
        )
    else:
        discrepancy_note = (
            f"The Data in Brief article reports {ARTICLE_REPORTED_COUNT:,} office "
            f"locations. The downloaded Excel file currently contains {clean_count:,} "
            "rows after cleaning. Validation findings are listed below for review; the "
            "script does not force the dataset to the article count without clear "
            "evidence that a row is invalid."
        )
    summary = f"""# Medicaid Office Cleaning Summary

- Input file path: `{RAW_INPUT.relative_to(PROJECT_ROOT)}`
- Output file path: `{CLEAN_OUTPUT.relative_to(PROJECT_ROOT)}`
- Available workbook sheets: {", ".join(f"`{sheet}`" for sheet in sheet_names)}
- Selected worksheet: `{selected_sheet}`
- Raw row count: {raw_count}
- Cleaned row count: {clean_count}
- Article-reported office count: {ARTICLE_REPORTED_COUNT}
- Rows removed for missing coordinates: {removed_missing_coordinates}
- States represented: {states_represented}
- Duplicate latitude/longitude pairs: {duplicate_coordinate_pairs}
- Exact duplicate source rows: {row_checks["duplicate_full_rows"]}
- Blank source rows: {row_checks["blank_rows"]}
- Rows with invalid or suspicious coordinates: {row_checks["invalid_coordinate_rows"]}
- Date/time run: {run_time}
- Raw file modified: No

## Workbook Structure

{sheet_note}

## Required Field Checks

{missing_required_text}

## Suspicious Row Checks

{suspicious_rows_text}

## Count Discrepancy Note

{discrepancy_note}

## Notes

The raw Excel file was not modified. Cleaning was performed by `scripts/02_clean_medicaid_offices.py`, and the cleaned CSV is generated locally under `data/processed/`.
"""
    SUMMARY_OUTPUT.write_text(summary, encoding="utf-8")


def main() -> int:
    if not RAW_INPUT.exists():
        raise FileNotFoundError(
            f"Missing raw input file: {RAW_INPUT.relative_to(PROJECT_ROOT)}. "
            "See docs/data_download_instructions.md."
        )

    headers, raw_rows, sheet_names, selected_sheet = read_workbook(RAW_INPUT)
    validate_expected_fields(headers)

    cleaned_rows: list[dict[str, str]] = []
    removed_missing_coordinates = 0

    for row in raw_rows:
        latitude = row.get("latitude", "")
        longitude = row.get("longitude", "")
        if not latitude or not longitude:
            removed_missing_coordinates += 1
            continue

        cleaned_row = {header: row.get(header, "") for header in headers}
        cleaned_row["full_address"] = build_full_address(cleaned_row)
        cleaned_row["office_id"] = build_office_id(cleaned_row)
        cleaned_rows.append(cleaned_row)

    state_counts = Counter(row.get("state", "") for row in cleaned_rows if row.get("state", ""))
    coordinate_counts = Counter(
        (row.get("latitude", ""), row.get("longitude", "")) for row in cleaned_rows
    )
    duplicate_coordinate_pairs = sum(1 for count in coordinate_counts.values() if count > 1)
    row_checks = summarize_row_checks(raw_rows)

    write_clean_csv(cleaned_rows, headers)
    write_summary(
        raw_count=len(raw_rows),
        clean_count=len(cleaned_rows),
        removed_missing_coordinates=removed_missing_coordinates,
        states_represented=len(state_counts),
        duplicate_coordinate_pairs=duplicate_coordinate_pairs,
        sheet_names=sheet_names,
        selected_sheet=selected_sheet,
        row_checks=row_checks,
    )

    print(f"Available workbook sheets: {', '.join(sheet_names)}")
    print(f"Selected worksheet: {selected_sheet}")
    print(f"Raw row count: {len(raw_rows)}")
    print(f"Cleaned row count: {len(cleaned_rows)}")
    print(f"Article-reported office count: {ARTICLE_REPORTED_COUNT}")
    print(f"Rows removed for missing coordinates: {removed_missing_coordinates}")
    print("Required field missing counts:")
    for field, count in row_checks["missing_required"].items():
        print(f"- {field}: {count}")
    print(f"Blank source rows: {row_checks['blank_rows']}")
    print(f"Exact duplicate source rows: {row_checks['duplicate_full_rows']}")
    print(f"Rows with invalid or suspicious coordinates: {row_checks['invalid_coordinate_rows']}")
    print(f"Suspicious metadata-like rows: {len(row_checks['suspicious_rows'])}")
    print(f"Number of states represented: {len(state_counts)}")
    print("Office count by state:")
    for state, count in sorted(state_counts.items()):
        print(f"- {state}: {count}")
    print(f"Duplicate latitude/longitude pairs: {duplicate_coordinate_pairs}")
    print(f"Wrote cleaned CSV: {CLEAN_OUTPUT.relative_to(PROJECT_ROOT)}")
    print(f"Wrote cleaning summary: {SUMMARY_OUTPUT.relative_to(PROJECT_ROOT)}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
