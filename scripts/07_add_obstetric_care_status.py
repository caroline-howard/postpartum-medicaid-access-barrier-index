"""Add county-level hospital-based obstetric care status.

Inputs:
    data/processed/county_office_access_with_acs_rurality.csv

Outputs:
    data/processed/hospital_obstetric_care_status.csv
    data/processed/county_postpartum_access_analytic_base.csv
    outputs/obstetric_care_status_summary.md

This step adds University of Minnesota Rural Health Research Center county
hospital-based obstetric care status only. It does not add Power BI files or
access index calculations.
"""

from __future__ import annotations

import csv
import re
import ssl
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen

import openpyxl


PROJECT_ROOT = Path(__file__).resolve().parents[1]
ANALYTIC_INPUT = PROJECT_ROOT / "data" / "processed" / "county_office_access_with_acs_rurality.csv"
MANUAL_OBSTETRIC_INPUT = PROJECT_ROOT / "data" / "manual" / "county_OBstat_2010_2024.xlsx"
OBSTETRIC_OUTPUT = PROJECT_ROOT / "data" / "processed" / "hospital_obstetric_care_status.csv"
MERGED_OUTPUT = PROJECT_ROOT / "data" / "processed" / "county_postpartum_access_analytic_base.csv"
SUMMARY_OUTPUT = PROJECT_ROOT / "outputs" / "obstetric_care_status_summary.md"

SOURCE_NAME = (
    "University of Minnesota Rural Health Research Center / Rural Maternal "
    "Health Data Support and Analysis Program"
)
DATASET_NAME = "2010-2024 County-Level Hospital-Based Obstetric Care Status"
SOURCE_PAGE = "https://rhrc.umn.edu/publication/2010-2024-county-level-hospital-based-obstetric-care-status/"
SOURCE_XLSX_URL = "https://rhrc.umn.edu/wp-content/uploads/2026/05/county_OBstat_2010_2024.xlsx"
SOURCE_SHEET_NAME = "county_OBstat_2010_2024"

# Keep this aligned with prior county base, ACS, and rurality steps.
ANALYTIC_STATE_FIPS_50_DC = {
    "01",
    "02",
    "04",
    "05",
    "06",
    "08",
    "09",
    "10",
    "11",
    "12",
    "13",
    "15",
    "16",
    "17",
    "18",
    "19",
    "20",
    "21",
    "22",
    "23",
    "24",
    "25",
    "26",
    "27",
    "28",
    "29",
    "30",
    "31",
    "32",
    "33",
    "34",
    "35",
    "36",
    "37",
    "38",
    "39",
    "40",
    "41",
    "42",
    "44",
    "45",
    "46",
    "47",
    "48",
    "49",
    "50",
    "51",
    "53",
    "54",
    "55",
    "56",
}

STATUS_LABEL_AVAILABLE = "Hospital-based obstetric care available"
STATUS_LABEL_NOT_IDENTIFIED = "No hospital-based obstetric care identified"

OBSTETRIC_BASE_FIELDS = [
    "county_fips",
    "state_abbr",
    "state_name",
    "obstetric_care_year",
    "has_hospital_obstetric_care",
    "no_hospital_obstetric_care",
    "obstetric_care_status_label",
    "has_hospital_obstetric_care_current",
]


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        raise FileNotFoundError(f"Required input file not found: {path}")
    with path.open(newline="", encoding="utf-8") as csv_file:
        return list(csv.DictReader(csv_file))


def write_csv_rows(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def download_or_load_workbook() -> tuple[Path, str, NamedTemporaryFile | None]:
    try:
        context = ssl.create_default_context()
        request = Request(
            SOURCE_XLSX_URL,
            headers={
                "User-Agent": (
                    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/125.0 Safari/537.36"
                )
            },
        )
        with urlopen(request, timeout=60, context=context) as response:
            content = response.read()
        temp_file = NamedTemporaryFile(delete=False, suffix=".xlsx")
        temp_file.write(content)
        temp_file.close()
        return Path(temp_file.name), SOURCE_XLSX_URL, temp_file
    except (HTTPError, URLError, TimeoutError, ssl.SSLError) as error:
        if MANUAL_OBSTETRIC_INPUT.exists():
            return MANUAL_OBSTETRIC_INPUT, str(MANUAL_OBSTETRIC_INPUT), None
        instructions = (
            "Unable to download the official hospital-based obstetric care status XLSX.\n"
            f"Source page: {SOURCE_PAGE}\n"
            f"XLSX URL: {SOURCE_XLSX_URL}\n"
            f"Manual fallback: download the XLSX and save it as {MANUAL_OBSTETRIC_INPUT}\n"
            f"Original error: {error}"
        )
        raise RuntimeError(instructions) from error


def normalize_header(value: Any) -> str:
    return str(value or "").strip()


def clean_county_fips(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, int):
        return str(value).zfill(5)
    text = str(value).strip()
    if not text:
        return ""
    if re.fullmatch(r"\d+(\.0)?", text):
        return str(int(float(text))).zfill(5)
    return text.zfill(5)


def clean_indicator(value: Any, *, field: str, county_fips: str) -> int | None:
    if value in {None, ""}:
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)) and value in {0, 1}:
        return int(value)
    text = str(value).strip().lower()
    if text in {"0", "0.0", "no", "n", "false"}:
        return 0
    if text in {"1", "1.0", "yes", "y", "true"}:
        return 1
    raise ValueError(f"Unexpected obstetric care value {value!r} in {field} for {county_fips}")


def read_obstetric_workbook(path: Path) -> tuple[list[dict[str, Any]], int, list[int]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if SOURCE_SHEET_NAME not in workbook.sheetnames:
        raise ValueError(
            f"Expected sheet {SOURCE_SHEET_NAME!r} not found. Available sheets: "
            + ", ".join(workbook.sheetnames)
        )
    sheet = workbook[SOURCE_SHEET_NAME]
    rows = sheet.iter_rows(values_only=True)
    try:
        header = [normalize_header(value) for value in next(rows)]
    except StopIteration as error:
        raise ValueError(f"Sheet {SOURCE_SHEET_NAME!r} is empty") from error

    required_fields = {"FIPS", "County_Name", "state"}
    missing_fields = sorted(required_fields - set(header))
    if missing_fields:
        raise ValueError("Obstetric care workbook is missing fields: " + ", ".join(missing_fields))

    year_fields = {
        int(match.group(1)): field
        for field in header
        if (match := re.fullmatch(r"AnyOB_(\d{4})", field))
    }
    if not year_fields:
        raise ValueError("No AnyOB_YYYY year fields found in obstetric care workbook")
    current_year = max(year_fields)

    output_rows: list[dict[str, Any]] = []
    for values in rows:
        source_row = dict(zip(header, values))
        county_fips = clean_county_fips(source_row.get("FIPS"))
        if not county_fips or county_fips[:2] not in ANALYTIC_STATE_FIPS_50_DC:
            continue

        year_values = {
            year: clean_indicator(source_row.get(field), field=field, county_fips=county_fips)
            for year, field in year_fields.items()
        }
        current_value = year_values[current_year]
        if current_value is None:
            raise ValueError(f"Missing current-year obstetric care value for {county_fips}")

        row: dict[str, Any] = {
            "county_fips": county_fips,
            "state_abbr": "",
            "state_name": str(source_row.get("state") or "").strip(),
            "obstetric_care_year": current_year,
            "has_hospital_obstetric_care": current_value,
            "no_hospital_obstetric_care": 1 - current_value,
            "obstetric_care_status_label": (
                STATUS_LABEL_AVAILABLE if current_value == 1 else STATUS_LABEL_NOT_IDENTIFIED
            ),
            "has_hospital_obstetric_care_current": current_value,
        }

        for year in sorted(year_fields):
            value = year_values[year]
            row[f"has_hospital_obstetric_care_{year}"] = "" if value is None else value

        if 2010 in year_values and year_values[2010] is not None:
            row["obstetric_care_loss_since_2010"] = (
                1 if year_values[2010] == 1 and current_value == 0 else 0
            )
        else:
            row["obstetric_care_loss_since_2010"] = ""

        output_rows.append(row)

    workbook.close()
    return output_rows, current_year, sorted(year_fields)


def enrich_with_analytic_fields(
    obstetric_rows: list[dict[str, Any]], analytic_rows: list[dict[str, str]]
) -> tuple[list[dict[str, Any]], list[dict[str, str]]]:
    obstetric_lookup = {row["county_fips"]: row for row in obstetric_rows}
    enriched_rows: list[dict[str, Any]] = []
    unmatched_rows: list[dict[str, str]] = []

    for analytic_row in sorted(analytic_rows, key=lambda row: row["county_fips"]):
        county_fips = analytic_row["county_fips"]
        if county_fips[:2] not in ANALYTIC_STATE_FIPS_50_DC:
            raise RuntimeError(
                "Analytic input contains a county outside the 50 states and D.C. "
                f"universe: {county_fips}"
            )
        source_row = obstetric_lookup.get(county_fips)
        if not source_row:
            unmatched_rows.append(analytic_row)
            continue

        row = dict(source_row)
        row["state_abbr"] = analytic_row.get("state_abbr", row.get("state_abbr", ""))
        row["state_name"] = analytic_row.get("state_name", row.get("state_name", ""))
        enriched_rows.append(row)

    return enriched_rows, unmatched_rows


def merge_obstetric_status(
    analytic_rows: list[dict[str, str]], obstetric_rows: list[dict[str, Any]]
) -> list[dict[str, Any]]:
    obstetric_lookup = {row["county_fips"]: row for row in obstetric_rows}
    merged_rows: list[dict[str, Any]] = []
    obstetric_fields = [
        field
        for field in obstetric_rows[0].keys()
        if field not in {"county_fips", "state_abbr", "state_name"}
    ]
    for analytic_row in analytic_rows:
        merged_row = dict(analytic_row)
        obstetric_row = obstetric_lookup.get(analytic_row["county_fips"], {})
        for field in obstetric_fields:
            merged_row[field] = obstetric_row.get(field, "")
        merged_rows.append(merged_row)
    return merged_rows


def to_int(value: str | int | None) -> int:
    if value in {None, ""}:
        return 0
    return int(value)


def summarize_by_field(
    merged_rows: list[dict[str, Any]], group_field: str
) -> list[dict[str, Any]]:
    summary: dict[str, dict[str, Any]] = defaultdict(
        lambda: {"county_count": 0, "with_obstetric_care": 0, "without_obstetric_care": 0}
    )
    for row in merged_rows:
        group_value = row.get(group_field, "Missing") or "Missing"
        current = to_int(row.get("has_hospital_obstetric_care_current"))
        summary[group_value]["county_count"] += 1
        summary[group_value]["with_obstetric_care"] += current
        summary[group_value]["without_obstetric_care"] += 1 - current

    rows: list[dict[str, Any]] = []
    for group_value in sorted(summary):
        values = summary[group_value]
        county_count = values["county_count"]
        without_count = values["without_obstetric_care"]
        rows.append(
            {
                group_field: group_value,
                "county_count": county_count,
                "with_obstetric_care": values["with_obstetric_care"],
                "without_obstetric_care": without_count,
                "share_without_obstetric_care": (
                    round(without_count / county_count, 6) if county_count else None
                ),
            }
        )
    return rows


def format_summary_rows(rows: list[dict[str, Any]], label_field: str) -> str:
    if not rows:
        return "- None"
    return "\n".join(
        "- {label}: {county_count} counties, {with_obstetric_care} with care, "
        "{without_obstetric_care} without care, {share_without_obstetric_care} share without care".format(
            label=row[label_field], **row
        )
        for row in rows
    )


def write_summary(
    *,
    source_location: str,
    analytic_count: int,
    matched_count: int,
    unmatched_count: int,
    states_represented: int,
    current_year: int,
    with_care: int,
    without_care: int,
    metro_summary_rows: list[dict[str, Any]],
    medicaid_office_summary_rows: list[dict[str, Any]],
) -> None:
    share_without = round(without_care / matched_count, 6) if matched_count else None
    run_timestamp = datetime.now().astimezone().isoformat(timespec="seconds")
    summary = f"""# Obstetric Care Status Summary

- Source: {SOURCE_NAME}
- Dataset: {DATASET_NAME}
- Source page: {SOURCE_PAGE}
- Source file: {source_location}
- Current obstetric care year used: {current_year}
- Analytic universe: 50 states and District of Columbia
- Input file: `{ANALYTIC_INPUT.relative_to(PROJECT_ROOT)}`
- Obstetric care output: `{OBSTETRIC_OUTPUT.relative_to(PROJECT_ROOT)}`
- Merged analytic output: `{MERGED_OUTPUT.relative_to(PROJECT_ROOT)}`
- Counties in input analytic file: {analytic_count}
- Counties with obstetric care status match: {matched_count}
- Counties missing obstetric care status: {unmatched_count}
- States/DC represented: {states_represented}
- Counties with hospital-based obstetric care: {with_care}
- Counties without hospital-based obstetric care: {without_care}
- Share of counties without hospital-based obstetric care: {share_without}
- Date/time run: {run_timestamp}

## Metro/Nonmetro Summary

{format_summary_rows(metro_summary_rows, "metro_nonmetro_flag")}

## Medicaid Office Availability Summary

{format_summary_rows(medicaid_office_summary_rows, "has_medicaid_office")}

## Notes on Limitations

This dataset identifies whether hospital-based obstetric care was available in each county and year. It does not capture all prenatal care, postpartum care, outpatient OB/GYN access, birth centers, doulas, midwives, community health centers, Medicaid enrollment support, or cross-county care seeking. This step adds clinical access context only and does not calculate the Potential Postpartum Medicaid Administrative Access Barrier Index.
"""
    SUMMARY_OUTPUT.write_text(summary, encoding="utf-8")


def main() -> int:
    analytic_rows = read_csv_rows(ANALYTIC_INPUT)
    workbook_path, source_location, temp_file = download_or_load_workbook()
    try:
        obstetric_rows, current_year, years = read_obstetric_workbook(workbook_path)
    finally:
        if temp_file is not None:
            Path(temp_file.name).unlink(missing_ok=True)

    enriched_obstetric_rows, unmatched_rows = enrich_with_analytic_fields(
        obstetric_rows, analytic_rows
    )
    merged_rows = merge_obstetric_status(analytic_rows, enriched_obstetric_rows)

    year_fields = [f"has_hospital_obstetric_care_{year}" for year in years]
    obstetric_fieldnames = [
        *OBSTETRIC_BASE_FIELDS,
        *year_fields,
        "obstetric_care_loss_since_2010",
    ]
    merged_fieldnames = list(dict.fromkeys([*analytic_rows[0].keys(), *obstetric_fieldnames]))

    write_csv_rows(OBSTETRIC_OUTPUT, enriched_obstetric_rows, obstetric_fieldnames)
    write_csv_rows(MERGED_OUTPUT, merged_rows, merged_fieldnames)

    matched_count = len(enriched_obstetric_rows)
    with_care = sum(to_int(row["has_hospital_obstetric_care_current"]) for row in enriched_obstetric_rows)
    without_care = matched_count - with_care
    states_represented = len({row["state_fips"] for row in merged_rows if row.get("state_fips")})
    metro_summary_rows = summarize_by_field(merged_rows, "metro_nonmetro_flag")
    medicaid_office_summary_rows = summarize_by_field(merged_rows, "has_medicaid_office")

    print(f"Counties in input analytic file: {len(analytic_rows)}")
    print(f"Counties with obstetric care status match: {matched_count}")
    print(f"Counties missing obstetric care status: {len(unmatched_rows)}")
    print(f"States/DC represented: {states_represented}")
    print(f"Current obstetric care year used: {current_year}")
    print(f"Counties with hospital-based obstetric care: {with_care}")
    print(f"Counties without hospital-based obstetric care: {without_care}")
    share_without = round(without_care / matched_count, 6) if matched_count else None
    print(f"Share of counties without hospital-based obstetric care: {share_without}")
    print("Breakdown by metro_nonmetro_flag:")
    for row in metro_summary_rows:
        print(
            "- {metro_nonmetro_flag}: {county_count} counties, {with_obstetric_care} with care, "
            "{without_obstetric_care} without care, {share_without_obstetric_care} share without care".format(
                **row
            )
        )
    print("Breakdown by has_medicaid_office:")
    for row in medicaid_office_summary_rows:
        print(
            "- {has_medicaid_office}: {county_count} counties, {with_obstetric_care} with care, "
            "{without_obstetric_care} without care, {share_without_obstetric_care} share without care".format(
                **row
            )
        )
    if unmatched_rows:
        print("Unmatched counties:")
        for row in unmatched_rows:
            print(f"- {row.get('county_name', '')} ({row.get('county_fips', '')})")

    write_summary(
        source_location=source_location,
        analytic_count=len(analytic_rows),
        matched_count=matched_count,
        unmatched_count=len(unmatched_rows),
        states_represented=states_represented,
        current_year=current_year,
        with_care=with_care,
        without_care=without_care,
        metro_summary_rows=metro_summary_rows,
        medicaid_office_summary_rows=medicaid_office_summary_rows,
    )

    print(f"Wrote obstetric care status: {OBSTETRIC_OUTPUT.relative_to(PROJECT_ROOT)}")
    print(f"Wrote merged analytic file: {MERGED_OUTPUT.relative_to(PROJECT_ROOT)}")
    print(f"Wrote summary: {SUMMARY_OUTPUT.relative_to(PROJECT_ROOT)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
